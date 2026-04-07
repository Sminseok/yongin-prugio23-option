import os
import sqlite3
from datetime import datetime
from functools import wraps
from pathlib import Path
from tempfile import NamedTemporaryFile

from flask import (
    Flask,
    flash,
    g,
    jsonify,
    redirect,
    render_template,
    request,
    send_from_directory,
    session,
    url_for,
)
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "data.db"
UPLOAD_DIR = BASE_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)
ALLOWED_EXTENSIONS = {".xlsx"}

OPTION_FIELDS = [
    "현관중문", "발코니확장", "쿡탑", "전기오븐", "식기세척기",
    "욕실 타일", "냉장고장", "욕실 복합환풍기", "붙박이장", "시스템에어컨",
    "시스템 청정환기", "조명특화", "84B 드레스룸", "복도벽", "벽지",
    "마루", "침실1 드레스룸", "침실1 드레스룸 제습기", "현관창고", "신발장",
    "주방벽, 상판", "주방수전", "주방 특화조명기구", "주방TV", "주방 팬트리",
    "주방가구", "샤워부스", "욕실 수전", "욕조 마감"
]

COLUMN_MAP = {
    "동": "dong",
    "층": "floor",
    "라인": "line",
    "호수": "hose",
    "평형": "unit_type",
    **{field: field for field in OPTION_FIELDS},
}

REQUIRED_HEADERS = list(COLUMN_MAP.keys())

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "change-this-secret-key")
app.config["MAX_CONTENT_LENGTH"] = 30 * 1024 * 1024
app.config["ADMIN_PASSWORD"] = os.environ.get("ADMIN_PASSWORD", "change-me")


def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(_error=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    db = sqlite3.connect(DB_PATH)
    db.executescript(
        """
        CREATE TABLE IF NOT EXISTS units (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dong TEXT NOT NULL,
            floor TEXT,
            line TEXT,
            hose TEXT NOT NULL,
            unit_type TEXT,
            "현관중문" TEXT,
            "발코니확장" TEXT,
            "쿡탑" TEXT,
            "전기오븐" TEXT,
            "식기세척기" TEXT,
            "욕실 타일" TEXT,
            "냉장고장" TEXT,
            "욕실 복합환풍기" TEXT,
            "붙박이장" TEXT,
            "시스템에어컨" TEXT,
            "시스템 청정환기" TEXT,
            "조명특화" TEXT,
            "84B 드레스룸" TEXT,
            "복도벽" TEXT,
            "벽지" TEXT,
            "마루" TEXT,
            "침실1 드레스룸" TEXT,
            "침실1 드레스룸 제습기" TEXT,
            "현관창고" TEXT,
            "신발장" TEXT,
            "주방벽, 상판" TEXT,
            "주방수전" TEXT,
            "주방 특화조명기구" TEXT,
            "주방TV" TEXT,
            "주방 팬트리" TEXT,
            "주방가구" TEXT,
            "샤워부스" TEXT,
            "욕실 수전" TEXT,
            "욕조 마감" TEXT,
            UNIQUE(dong, hose)
        );

        CREATE INDEX IF NOT EXISTS idx_units_dong_hose ON units(dong, hose);

        CREATE TABLE IF NOT EXISTS uploads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT NOT NULL,
            uploaded_at TEXT NOT NULL,
            row_count INTEGER NOT NULL
        );
        """
    )
    db.commit()
    db.close()


def allowed_file(filename: str) -> bool:
    return Path(filename).suffix.lower() in ALLOWED_EXTENSIONS


def admin_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if not session.get("is_admin"):
            return redirect(url_for("admin_login", next=request.path))
        return view_func(*args, **kwargs)
    return wrapped


def normalize_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def parse_excel(file_path: Path, sheet_name: str = "옵션현황"):
    wb = load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"'{sheet_name}' 시트를 찾을 수 없습니다.")

    ws = wb[sheet_name]
    headers = [normalize_value(c.value) for c in ws[1]]

    missing = [h for h in REQUIRED_HEADERS if h not in headers]
    if missing:
        raise ValueError("필수 컬럼이 없습니다: " + ", ".join(missing))

    header_index = {h: i for i, h in enumerate(headers)}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue
        record = {}
        for source_col, target_col in COLUMN_MAP.items():
            idx = header_index[source_col]
            record[target_col] = normalize_value(row[idx] if idx < len(row) else None)
        if not record["dong"] or not record["hose"]:
            continue
        rows.append(record)
    return rows


def replace_units(rows, filename="initial.xlsx"):
    db = sqlite3.connect(DB_PATH)
    cursor = db.cursor()
    cursor.execute("DELETE FROM units")

    sql = """
        INSERT INTO units (
            dong, floor, line, hose, unit_type,
            "현관중문", "발코니확장", "쿡탑", "전기오븐", "식기세척기",
            "욕실 타일", "냉장고장", "욕실 복합환풍기", "붙박이장", "시스템에어컨",
            "시스템 청정환기", "조명특화", "84B 드레스룸", "복도벽", "벽지",
            "마루", "침실1 드레스룸", "침실1 드레스룸 제습기", "현관창고", "신발장",
            "주방벽, 상판", "주방수전", "주방 특화조명기구", "주방TV", "주방 팬트리",
            "주방가구", "샤워부스", "욕실 수전", "욕조 마감"
        ) VALUES (
            ?, ?, ?, ?, ?,
            ?, ?, ?, ?, ?,
            ?, ?, ?, ?, ?,
            ?, ?, ?, ?, ?,
            ?, ?, ?, ?, ?,
            ?, ?, ?, ?, ?,
            ?, ?, ?, ?
        )
    """

    for row in rows:
        cursor.execute(
            sql,
            [
                row["dong"], row["floor"], row["line"], row["hose"], row["unit_type"],
                row["현관중문"], row["발코니확장"], row["쿡탑"], row["전기오븐"], row["식기세척기"],
                row["욕실 타일"], row["냉장고장"], row["욕실 복합환풍기"], row["붙박이장"], row["시스템에어컨"],
                row["시스템 청정환기"], row["조명특화"], row["84B 드레스룸"], row["복도벽"], row["벽지"],
                row["마루"], row["침실1 드레스룸"], row["침실1 드레스룸 제습기"], row["현관창고"], row["신발장"],
                row["주방벽, 상판"], row["주방수전"], row["주방 특화조명기구"], row["주방TV"], row["주방 팬트리"],
                row["주방가구"], row["샤워부스"], row["욕실 수전"], row["욕조 마감"],
            ],
        )

    cursor.execute(
        "INSERT INTO uploads (filename, uploaded_at, row_count) VALUES (?, ?, ?)",
        (filename, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), len(rows)),
    )
    db.commit()
    db.close()


def get_stats():
    db = get_db()
    total_units = db.execute("SELECT COUNT(*) AS cnt FROM units").fetchone()["cnt"]
    total_dongs = db.execute("SELECT COUNT(DISTINCT dong) AS cnt FROM units").fetchone()["cnt"]
    last_upload = db.execute(
        "SELECT filename, uploaded_at, row_count FROM uploads ORDER BY id DESC LIMIT 1"
    ).fetchone()
    return {
        "total_units": total_units,
        "total_dongs": total_dongs,
        "last_upload": dict(last_upload) if last_upload else None,
    }


@app.route("/")
def index():
    return render_template("index.html", option_fields=OPTION_FIELDS)


@app.route("/api/summary")
def api_summary():
    db = get_db()
    dongs = [row["dong"] for row in db.execute("SELECT DISTINCT dong FROM units ORDER BY CAST(dong AS INTEGER), dong").fetchall()]
    return jsonify({**get_stats(), "dongs": dongs})


@app.route("/api/hoses")
def api_hoses():
    dong = request.args.get("dong", "").strip()
    if not dong:
        return jsonify({"hoses": []})
    db = get_db()
    rows = db.execute(
        "SELECT hose FROM units WHERE dong = ? ORDER BY CAST(hose AS INTEGER), hose", (dong,)
    ).fetchall()
    return jsonify({"hoses": [r["hose"] for r in rows]})


@app.route("/api/unit")
def api_unit():
    dong = request.args.get("dong", "").strip()
    hose = request.args.get("hose", "").strip()
    if not dong or not hose:
        return jsonify({"error": "동과 호수를 모두 입력해 주세요."}), 400

    db = get_db()
    row = db.execute("SELECT * FROM units WHERE dong = ? AND hose = ?", (dong, hose)).fetchone()
    if not row:
        return jsonify({"error": "해당 세대 데이터를 찾지 못했습니다."}), 404

    data = dict(row)
    options = {field: data.pop(field, "") for field in OPTION_FIELDS}
    data["options"] = options
    return jsonify(data)


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        password = request.form.get("password", "")
        if password == app.config["ADMIN_PASSWORD"]:
            session["is_admin"] = True
            flash("관리자 로그인 완료", "success")
            return redirect(request.args.get("next") or url_for("admin"))
        flash("비밀번호가 올바르지 않습니다.", "error")
    return render_template("admin_login.html")


@app.route("/admin/logout")
def admin_logout():
    session.clear()
    flash("로그아웃되었습니다.", "success")
    return redirect(url_for("admin_login"))


@app.route("/admin", methods=["GET", "POST"])
@admin_required
def admin():
    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename:
            flash("업로드할 엑셀 파일을 선택해 주세요.", "error")
            return redirect(url_for("admin"))

        if not allowed_file(file.filename):
            flash(".xlsx 파일만 업로드할 수 있습니다.", "error")
            return redirect(url_for("admin"))

        filename = secure_filename(file.filename)
        temp_path = None
        try:
            with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                temp_path = Path(tmp.name)
                file.save(temp_path)

            rows = parse_excel(temp_path)
            replace_units(rows, filename=filename)

            saved_path = UPLOAD_DIR / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
            temp_path.replace(saved_path)
            flash(f"업로드 완료: {len(rows)}개 세대 데이터 반영", "success")
        except Exception as e:
            if temp_path and temp_path.exists():
                temp_path.unlink(missing_ok=True)
            flash(f"업로드 실패: {e}", "error")
        return redirect(url_for("admin"))

    db = get_db()
    upload_history = db.execute(
        "SELECT filename, uploaded_at, row_count FROM uploads ORDER BY id DESC LIMIT 10"
    ).fetchall()
    return render_template("admin.html", stats=get_stats(), history=upload_history)


@app.route("/health")
def health():
    return {"status": "ok"}


@app.route("/uploads/<path:filename>")
@admin_required
def uploaded_file(filename):
    return send_from_directory(UPLOAD_DIR, filename)


def seed_if_empty():
    if not DB_PATH.exists():
        init_db()
    else:
        init_db()

    db = sqlite3.connect(DB_PATH)
    count = db.execute("SELECT COUNT(*) FROM units").fetchone()[0]
    db.close()
    if count == 0:
        sample = BASE_DIR / "seed.xlsx"
        if sample.exists():
            rows = parse_excel(sample)
            replace_units(rows, filename=sample.name)


if __name__ == "__main__":
    seed_if_empty()
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True)
